#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to convert Algerian Education Ministry exam lists (Markdown format) to Excel
Supports Arabic RTL layout, committee-based sheets, and professional formatting

Usage:
    python md_to_excel.py input.md
    python md_to_excel.py input.md -o output_name
    python md_to_excel.py input.md --single-only
    python md_to_excel.py input.md --multi-only
"""

import re
import sys
import os
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def parse_md_file(md_file_path):
    """
    Parse the Markdown file and extract student records.

    Expected format per line:
    ## 47112001 [birthplace] DD-MM-YYYY [name parts...] 2947XXXX [order]
    """
    with open(md_file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.splitlines()

    all_records = []
    current_committee = None
    current_page = None
    center_info = None  # معلومات مركز الإجراء
    branch_info = None  # معلومات الفرع
    wilaya_info = None  # معلومات الولاية

    for line in lines:
        line = line.strip()

        # Extract branch info: ## غرداية الفـــــرع:
        if 'الفـــــرع:' in line or 'الفرع:' in line:
            match = re.search(r'##\s*(.+?)\s+الف', line)
            if match:
                branch_info = match.group(1).strip()

        # Extract wilaya info: ## المنيعة الولايــــة:
        if 'الولايــــة:' in line or 'الولاية:' in line:
            match = re.search(r'##\s*(.+?)\s+الولاي', line)
            if match:
                wilaya_info = match.group(1).strip()

        # Extract center info: ## 58003/ثانوية الشيخ محمد بلكبير - المنيعة مركز الإجراء:
        if 'مركز الإجراء:' in line and '##' in line:
            # استخراج النص بعد ##
            match = re.search(r'##\s*(.+?)\s+مركز الإجراء:', line)
            if match:
                center_info = match.group(1).strip()

        # Extract committee number: # 94748 لجنة التلميذ
        if 'لجنة التلميذ' in line and '#' in line:
            match = re.search(r'#\s*(\d+)', line)
            if match:
                current_committee = match.group(1)

        # Extract page info: صفحة X من Y
        if 'صفحة' in line and 'من' in line:
            match = re.search(r'صفحة\s+(\d+)\s+من\s+(\d+)', line)
            if match:
                current_page = match.group(1)

        # Extract student records: starts with ## and ends with registration number + order
        # Pattern: ## 47112001 ... DD-MM-YYYY ... 2947XXXX X
        if line.startswith('## ') and re.search(r'\d{8}\s+\d+$', line):
            data = line[2:].strip()
            parts = data.split()

            if len(parts) < 6:
                continue

            institution = parts[0]

            # Find registration number (8 digits, starts with 2947)
            reg_num = ''
            order = ''
            reg_idx = -1

            for i, part in enumerate(parts):
                # البحث عن رقم تسجيل من 8 أرقام يبدأ بـ 29
                if re.match(r'^\d{8}$', part) and part.startswith('29'):
                    reg_num = part
                    reg_idx = i
                    if i + 1 < len(parts) and parts[i + 1].isdigit():
                        order = parts[i + 1]
                    break

            if not reg_num or not order or reg_idx <= 0:
                continue

            # Find birth date (DD-MM-YYYY)
            date = ''
            date_idx = -1
            for i, part in enumerate(parts):
                if re.match(r'^\d{2}-\d{2}-\d{4}$', part):
                    date = part
                    date_idx = i
                    break

            if not date or date_idx <= 1:
                continue

            # Birthplace: between institution and date
            birthplace = ' '.join(parts[1:date_idx])

            # Name and surname: between date and registration number
            name_parts = parts[date_idx + 1:reg_idx]

            if len(name_parts) >= 2:
                surname = name_parts[-1]
                name = ' '.join(name_parts[:-1])
            elif len(name_parts) == 1:
                name = name_parts[0]
                surname = ''
            else:
                continue

            all_records.append({
                'committee': current_committee or '',
                'page': current_page or '',
                'institution': institution,
                'birthplace': birthplace,
                'birth_date': date,
                'name': name,
                'surname': surname,
                'reg_num': reg_num,
                'order': int(order),
                'center_info': center_info,
                'branch_info': branch_info,
                'wilaya_info': wilaya_info
            })

    return all_records


def create_styles():
    """Create and return all Excel styles."""
    styles = {
        'header_font': Font(name='Arial', size=11, bold=True, color="FFFFFF"),
        'header_fill': PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),

        'title_font': Font(name='Arial', size=14, bold=True, color="1F4E78"),
        'subtitle_font': Font(name='Arial', size=11, bold=True, color="1F4E78"),

        'info_font': Font(name='Arial', size=10),
        'info_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),

        'light_fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        'white_fill': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),

        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    return styles


def setup_sheet_headers(ws, styles, committee_num, total_students, center_info=None, branch_info=None, wilaya_info=None):
    """Setup the header section of a worksheet."""
    # RTL direction
    ws.sheet_view.rightToLeft = True

    # Title rows
    ws.merge_cells('A1:H1')
    ws['A1'] = "الجمهورية الجزائرية الديمقراطية الشعبية"
    ws['A1'].font = Font(name='Arial', size=12, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:H2')
    ws['A2'] = "وزارة التربية الوطنية"
    ws['A2'].font = Font(name='Arial', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    ws.merge_cells('A3:H3')
    ws['A3'] = "الديوان الوطني للامتحانات والمسابقات"
    ws['A3'].font = Font(name='Arial', size=12, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25

    ws.merge_cells('A4:H4')
    ws['A4'] = "امتحان: شهادة التعليم المتوسط - دورة 2026"
    ws['A4'].font = Font(name='Arial', size=12, bold=True, color="1F4E78")
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 25

    ws.merge_cells('A5:H5')
    ws['A5'] = "قوائم المناداة حسب لجان التلاميذ"
    ws['A5'].font = Font(name='Arial', size=13, bold=True, color="1F4E78")
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 30

    # Info row - استخدام المعلومات الديناميكية
    ws.merge_cells('A6:H6')
    if branch_info and wilaya_info and center_info:
        ws['A6'] = f"الفرع: {branch_info} | الولاية: {wilaya_info} | مركز الإجراء: {center_info}"
    else:
        ws['A6'] = "الفرع: غرداية | الولاية: المنيعة | مركز الإجراء: 58003/ثانوية الشيخ محمد بلكبير - المنيعة"
    ws['A6'].font = Font(name='Arial', size=10, bold=True)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[6].height = 30

    # Committee info
    ws.merge_cells('A7:H7')
    ws['A7'] = f"لجنة التلميذ: {committee_num} | عدد التلاميذ: {total_students}"
    ws['A7'].font = Font(name='Arial', size=11, bold=True, color="C00000")
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[7].height = 25

    # Empty row
    ws.row_dimensions[8].height = 10

    # Column headers - إضافة عمود الجنس بعد تاريخ الميلاد
    headers = ["ترتيب", "ر.التسجيل", "اللقب", "الاسم", "تا. الميلاد", "الجنس", "م.الميلاد", "المؤسسة"]
    header_row = 9

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']

    ws.row_dimensions[header_row].height = 30

    # Column widths
    ws.column_dimensions['A'].width = 8   # ترتيب
    ws.column_dimensions['B'].width = 14  # ر.التسجيل
    ws.column_dimensions['C'].width = 18  # اللقب
    ws.column_dimensions['D'].width = 24  # الاسم
    ws.column_dimensions['E'].width = 14  # تا. الميلاد
    ws.column_dimensions['F'].width = 8   # الجنس (عمود فارغ)
    ws.column_dimensions['G'].width = 16  # م.الميلاد
    ws.column_dimensions['H'].width = 14  # المؤسسة

    return header_row


def write_data_rows(ws, styles, records, start_row=10):
    """Write student data rows to the worksheet."""
    for i, record in enumerate(records):
        row = start_row + i
        row_fill = styles['light_fill'] if i % 2 == 0 else styles['white_fill']

        # إضافة عمود فارغ للجنس بعد تاريخ الميلاد
        cells_data = [
            record['order'],
            record['reg_num'],
            record['surname'],
            record['name'],
            record['birth_date'],
            '',  # عمود الجنس (فارغ)
            record['birthplace'],
            record['institution']
        ]

        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.font = styles['info_font']
            cell.alignment = styles['info_alignment']
            cell.border = styles['thin_border']
            cell.fill = row_fill

        ws.row_dimensions[row].height = 22

    return start_row + len(records)


def add_footer(ws, last_row, records_count):
    """Add footer row with summary info."""
    footer_row = last_row + 1
    ws.merge_cells(f'A{footer_row}:H{footer_row}')
    ws[f'A{footer_row}'] = f"عدد التلاميذ: {records_count} | تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws[f'A{footer_row}'].font = Font(name='Arial', size=9, italic=True)
    ws[f'A{footer_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[footer_row].height = 20
    return footer_row


def create_single_sheet_excel(records, output_path):
    """Create a single-sheet Excel file with all records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "جميع التلاميذ"

    styles = create_styles()

    # الحصول على معلومات المركز من أول سجل
    center_info = records[0].get('center_info') if records else None
    branch_info = records[0].get('branch_info') if records else None
    wilaya_info = records[0].get('wilaya_info') if records else None

    # Setup headers (use "جميع اللجان" as committee number)
    header_row = setup_sheet_headers(ws, styles, "جميع اللجان", len(records), center_info, branch_info, wilaya_info)

    # Write data
    last_data_row = write_data_rows(ws, styles, records, start_row=header_row + 1)

    # Add footer
    add_footer(ws, last_data_row, len(records))

    # Freeze panes
    ws.freeze_panes = f'A{header_row + 1}'

    wb.save(output_path)
    print(f"[OK] Single-sheet Excel: {output_path}")
    print(f"     Total records: {len(records)}")


def create_multi_sheet_excel(records, output_path):
    """Create a multi-sheet Excel file with one sheet per committee."""
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    styles = create_styles()

    # Get unique committees sorted
    committees = sorted(set(r['committee'] for r in records))

    for committee in committees:
        committee_records = [r for r in records if r['committee'] == committee]

        ws = wb.create_sheet(title=f"لجنة_{committee}")

        # الحصول على معلومات المركز من أول سجل في اللجنة
        center_info = committee_records[0].get('center_info') if committee_records else None
        branch_info = committee_records[0].get('branch_info') if committee_records else None
        wilaya_info = committee_records[0].get('wilaya_info') if committee_records else None

        # Setup headers
        header_row = setup_sheet_headers(ws, styles, committee, len(committee_records), center_info, branch_info, wilaya_info)

        # Write data
        last_data_row = write_data_rows(ws, styles, committee_records, start_row=header_row + 1)

        # Add footer
        add_footer(ws, last_data_row, len(committee_records))

        # Freeze panes
        ws.freeze_panes = f'A{header_row + 1}'

    wb.save(output_path)
    print(f"[OK] Multi-sheet Excel: {output_path}")
    print(f"     Total committees: {len(committees)}")
    print(f"     Total records: {len(records)}")
    for committee in committees:
        count = len([r for r in records if r['committee'] == committee])
        print(f"     - Committee {committee}: {count} students")


def main():
    parser = argparse.ArgumentParser(
        description='Convert Algerian exam lists from Markdown to Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python md_to_excel.py input.md                    # Creates both files in current directory
  python md_to_excel.py input.md -o results/output  # Custom output path
  python md_to_excel.py input.md --single-only      # Only single-sheet
  python md_to_excel.py input.md --multi-only       # Only multi-sheet
        """
    )

    parser.add_argument('input_file', help='Input Markdown file path')
    parser.add_argument('-o', '--output', help='Output path/name (without extension)')
    parser.add_argument('--single-only', action='store_true', help='Only create single-sheet file')
    parser.add_argument('--multi-only', action='store_true', help='Only create multi-sheet file')

    args = parser.parse_args()

    # Validate input file
    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"[ERROR] File not found: {input_path}")
        sys.exit(1)

    # Parse the file
    print(f"[*] Parsing {input_path}...")
    records = parse_md_file(str(input_path))

    if not records:
        print("[ERROR] No records found in the file.")
        print("        Make sure the file follows the expected format:")
        print("        ## 47112001 [birthplace] DD-MM-YYYY [name] [surname] 2947XXXX [order]")
        sys.exit(1)

    print(f"[*] Found {len(records)} student records")

    # Determine output path
    if args.output:
        output_path = Path(args.output)
        # Create parent directories if they don't exist
        output_path.parent.mkdir(parents=True, exist_ok=True)
        base_name = str(output_path)
    else:
        # Use current working directory
        base_name = str(Path.cwd() / input_path.stem)

    # Generate files based on options
    if not args.multi_only:
        single_output = f"{base_name}_single.xlsx"
        create_single_sheet_excel(records, single_output)

    if not args.single_only:
        multi_output = f"{base_name}_committees.xlsx"
        create_multi_sheet_excel(records, multi_output)

    print("[*] Done!")


if __name__ == "__main__":
    main()