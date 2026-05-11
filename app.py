#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Web-based Script to convert Algerian Education Ministry exam lists (Markdown format) to Excel
Supports Arabic RTL layout, committee-based sheets, and professional formatting

Usage:
    python md_to_excel_web.py                        # Start web server on port 5000
    python md_to_excel_web.py --port 8080            # Custom port
    
    OR use CLI mode:
    python md_to_excel_web.py --cli input.md         # CLI mode
    python md_to_excel_web.py --cli input.md -o output_name
"""

import re
import sys
import os
import argparse
from datetime import datetime
from pathlib import Path
from io import BytesIO
import tempfile

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

try:
    from flask import Flask, render_template_string, request, send_file, jsonify
    FLASK_AVAILABLE = True
except ImportError:
    FLASK_AVAILABLE = False
    print("Warning: Flask not installed. Web mode disabled. Install with: pip install flask")


def parse_md_content(content):
    """
    Parse the Markdown content and extract student records.

    Expected format per line:
    ## 47112001 [birthplace] DD-MM-YYYY [name parts...] 2947XXXX [order]
    """
    lines = content.splitlines()

    all_records = []
    current_committee = None
    current_page = None
    center_info = None  # معلومات مركز الإجراء
    branch_info = None  # معلومات الفرع
    wilaya_info = None  # معلومات الولاية

    for line in lines:
        line = line.strip()

        # Extract branch info: ## غرداية الفـــــرع:
        if 'الفـــــرع:' in line or 'الفرع:' in line:
            match = re.search(r'##\s*(.+?)\s+الف', line)
            if match:
                branch_info = match.group(1).strip()

        # Extract wilaya info: ## المنيعة الولايــــة:
        if 'الولايــــة:' in line or 'الولاية:' in line:
            match = re.search(r'##\s*(.+?)\s+الولاي', line)
            if match:
                wilaya_info = match.group(1).strip()

        # Extract center info: ## 58003/ثانوية الشيخ محمد بلكبير - المنيعة مركز الإجراء:
        if 'مركز الإجراء:' in line and '##' in line:
            # استخراج النص بعد ##
            match = re.search(r'##\s*(.+?)\s+مركز الإجراء:', line)
            if match:
                center_info = match.group(1).strip()

        # Extract committee number: # 94748 لجنة التلميذ
        if 'لجنة التلميذ' in line and '#' in line:
            match = re.search(r'#\s*(\d+)', line)
            if match:
                current_committee = match.group(1)

        # Extract page info: صفحة X من Y
        if 'صفحة' in line and 'من' in line:
            match = re.search(r'صفحة\s+(\d+)\s+من\s+(\d+)', line)
            if match:
                current_page = match.group(1)

        # Extract student records: starts with ## and ends with registration number + order
        # Pattern: ## 47112001 ... DD-MM-YYYY ... 2947XXXX X
        if line.startswith('## ') and re.search(r'\d{8}\s+\d+$', line):
            data = line[2:].strip()
            parts = data.split()

            if len(parts) < 6:
                continue

            institution = parts[0]

            # Find registration number (8 digits, starts with 2947)
            reg_num = ''
            order = ''
            reg_idx = -1

            for i, part in enumerate(parts):
                # البحث عن رقم تسجيل من 8 أرقام يبدأ بـ 29
                if re.match(r'^\d{8}$', part) and part.startswith('29'):
                    reg_num = part
                    reg_idx = i
                    if i + 1 < len(parts) and parts[i + 1].isdigit():
                        order = parts[i + 1]
                    break

            if not reg_num or not order or reg_idx <= 0:
                continue

            # Find birth date (DD-MM-YYYY)
            date = ''
            date_idx = -1
            for i, part in enumerate(parts):
                if re.match(r'^\d{2}-\d{2}-\d{4}$', part):
                    date = part
                    date_idx = i
                    break

            if not date or date_idx <= 1:
                continue

            # Birthplace: between institution and date
            birthplace = ' '.join(parts[1:date_idx])

            # Name and surname: between date and registration number
            name_parts = parts[date_idx + 1:reg_idx]

            if len(name_parts) >= 2:
                surname = name_parts[-1]
                name = ' '.join(name_parts[:-1])
            elif len(name_parts) == 1:
                name = name_parts[0]
                surname = ''
            else:
                continue

            all_records.append({
                'committee': current_committee or '',
                'page': current_page or '',
                'institution': institution,
                'birthplace': birthplace,
                'birth_date': date,
                'name': name,
                'surname': surname,
                'reg_num': reg_num,
                'order': int(order),
                'center_info': center_info,
                'branch_info': branch_info,
                'wilaya_info': wilaya_info
            })

    return all_records


def parse_md_file(md_file_path):
    """Parse the Markdown file and extract student records."""
    with open(md_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    return parse_md_content(content)

def create_styles():
    """Create and return all Excel styles."""
    styles = {
        'header_font': Font(name='Arial', size=11, bold=True, color="FFFFFF"),
        'header_fill': PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),

        'title_font': Font(name='Arial', size=14, bold=True, color="1F4E78"),
        'subtitle_font': Font(name='Arial', size=11, bold=True, color="1F4E78"),

        'info_font': Font(name='Arial', size=10),
        'info_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),

        'light_fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        'white_fill': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),

        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    return styles

def setup_sheet_headers(ws, styles, committee_num, total_students, center_info=None, branch_info=None, wilaya_info=None):
    """Setup the header section of a worksheet."""
    # RTL direction
    ws.sheet_view.rightToLeft = True

    # Title rows
    ws.merge_cells('A1:H1')
    ws['A1'] = "الجمهورية الجزائرية الديمقراطية الشعبية"
    ws['A1'].font = Font(name='Arial', size=12, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:H2')
    ws['A2'] = "وزارة التربية الوطنية"
    ws['A2'].font = Font(name='Arial', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    ws.merge_cells('A3:H3')
    ws['A3'] = "الديوان الوطني للامتحانات والمسابقات"
    ws['A3'].font = Font(name='Arial', size=12, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25

    ws.merge_cells('A4:H4')
    ws['A4'] = "امتحان: شهادة التعليم المتوسط - دورة 2026"
    ws['A4'].font = Font(name='Arial', size=12, bold=True, color="1F4E78")
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 25

    ws.merge_cells('A5:H5')
    ws['A5'] = "قوائم المناداة حسب لجان التلاميذ"
    ws['A5'].font = Font(name='Arial', size=13, bold=True, color="1F4E78")
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 30

    # Info row - استخدام المعلومات الديناميكية
    ws.merge_cells('A6:H6')
    if branch_info and wilaya_info and center_info:
        ws['A6'] = f"الفرع: {branch_info} | الولاية: {wilaya_info} | مركز الإجراء: {center_info}"
    else:
        ws['A6'] = "الفرع: غرداية | الولاية: المنيعة | مركز الإجراء: 58003/ثانوية الشيخ محمد بلكبير - المنيعة"
    ws['A6'].font = Font(name='Arial', size=10, bold=True)
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[6].height = 30

    # Committee info
    ws.merge_cells('A7:H7')
    ws['A7'] = f"لجنة التلميذ: {committee_num} | عدد التلاميذ: {total_students}"
    ws['A7'].font = Font(name='Arial', size=11, bold=True, color="C00000")
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[7].height = 25

    # Empty row
    ws.row_dimensions[8].height = 10

    # Column headers - إضافة عمود الجنس بعد تاريخ الميلاد
    headers = ["ترتيب", "ر.التسجيل", "اللقب", "الاسم", "تا. الميلاد", "الجنس", "م.الميلاد", "المؤسسة"]
    header_row = 9

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']

    ws.row_dimensions[header_row].height = 30

    # Column widths
    ws.column_dimensions['A'].width = 8   # ترتيب
    ws.column_dimensions['B'].width = 14  # ر.التسجيل
    ws.column_dimensions['C'].width = 18  # اللقب
    ws.column_dimensions['D'].width = 24  # الاسم
    ws.column_dimensions['E'].width = 14  # تا. الميلاد
    ws.column_dimensions['F'].width = 8   # الجنس (عمود فارغ)
    ws.column_dimensions['G'].width = 16  # م.الميلاد
    ws.column_dimensions['H'].width = 14  # المؤسسة

    return header_row


def write_data_rows(ws, styles, records, start_row=10):
    """Write student data rows to the worksheet."""
    for i, record in enumerate(records):
        row = start_row + i
        row_fill = styles['light_fill'] if i % 2 == 0 else styles['white_fill']

        # إضافة عمود فارغ للجنس بعد تاريخ الميلاد
        cells_data = [
            record['order'],
            record['reg_num'],
            record['surname'],
            record['name'],
            record['birth_date'],
            '',  # عمود الجنس (فارغ)
            record['birthplace'],
            record['institution']
        ]

        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.font = styles['info_font']
            cell.alignment = styles['info_alignment']
            cell.border = styles['thin_border']
            cell.fill = row_fill

        ws.row_dimensions[row].height = 22

    return start_row + len(records)


def add_footer(ws, last_row, records_count):
    """Add footer row with summary info."""
    footer_row = last_row + 1
    ws.merge_cells(f'A{footer_row}:H{footer_row}')
    ws[f'A{footer_row}'] = f"عدد التلاميذ: {records_count} | تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws[f'A{footer_row}'].font = Font(name='Arial', size=9, italic=True)
    ws[f'A{footer_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[footer_row].height = 20
    return footer_row


def create_single_sheet_excel(records, output_path=None):
    """Create a single-sheet Excel file with all records. Returns BytesIO if no output_path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "جميع التلاميذ"

    styles = create_styles()

    # الحصول على معلومات المركز من أول سجل
    center_info = records[0].get('center_info') if records else None
    branch_info = records[0].get('branch_info') if records else None
    wilaya_info = records[0].get('wilaya_info') if records else None

    # Setup headers (use "جميع اللجان" as committee number)
    header_row = setup_sheet_headers(ws, styles, "جميع اللجان", len(records), center_info, branch_info, wilaya_info)

    # Write data
    last_data_row = write_data_rows(ws, styles, records, start_row=header_row + 1)

    # Add footer
    add_footer(ws, last_data_row, len(records))

    # Freeze panes
    ws.freeze_panes = f'A{header_row + 1}'

    if output_path:
        wb.save(output_path)
        print(f"[OK] Single-sheet Excel: {output_path}")
        print(f"     Total records: {len(records)}")
        return output_path
    else:
        # Return BytesIO for web mode
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def create_multi_sheet_excel(records, output_path=None):
    """Create a multi-sheet Excel file with one sheet per committee. Returns BytesIO if no output_path."""
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    styles = create_styles()

    # Get unique committees sorted
    committees = sorted(set(r['committee'] for r in records))

    for committee in committees:
        committee_records = [r for r in records if r['committee'] == committee]

        ws = wb.create_sheet(title=f"لجنة_{committee}")

        # الحصول على معلومات المركز من أول سجل في اللجنة
        center_info = committee_records[0].get('center_info') if committee_records else None
        branch_info = committee_records[0].get('branch_info') if committee_records else None
        wilaya_info = committee_records[0].get('wilaya_info') if committee_records else None

        # Setup headers
        header_row = setup_sheet_headers(ws, styles, committee, len(committee_records), center_info, branch_info, wilaya_info)

        # Write data
        last_data_row = write_data_rows(ws, styles, committee_records, start_row=header_row + 1)

        # Add footer
        add_footer(ws, last_data_row, len(committee_records))

        # Freeze panes
        ws.freeze_panes = f'A{header_row + 1}'

    if output_path:
        wb.save(output_path)
        print(f"[OK] Multi-sheet Excel: {output_path}")
        print(f"     Total committees: {len(committees)}")
        print(f"     Total records: {len(records)}")
        for committee in committees:
            count = len([r for r in records if r['committee'] == committee])
            print(f"     - Committee {committee}: {count} students")
        return output_path
    else:
        # Return BytesIO for web mode
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# HTML Template
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>محول قوائم الامتحانات - وزارة التربية الوطنية</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        body { direction: rtl; }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        .spinner {
            animation: spin 1s linear infinite;
        }
    </style>
</head>
<body class="bg-gradient-to-br from-blue-400 via-purple-400 to-purple-700 min-h-screen p-5 text-right">
    <div class="max-w-2xl mx-auto bg-white rounded-3xl shadow-2xl overflow-hidden">
        <!-- Flag Bar -->
        <div class="h-2 bg-gradient-to-r from-green-700 to-white"></div>
        
        <!-- Header -->
        <div class="bg-gradient-to-br from-green-700 to-green-900 text-white p-8 text-center">
            <h1 class="text-4xl font-bold mb-2">🇩🇿 محول قوائم الامتحانات</h1>
            <p class="text-lg opacity-90">وزارة التربية الوطنية - الجمهورية الجزائرية الديمقراطية الشعبية</p>
        </div>
        
        <!-- Content -->
        <div class="p-10">
            <form id="converterForm" method="POST" class="space-y-6">
                <!-- Input Group -->
                <div>
                    <label class="block text-lg font-semibold text-gray-800 mb-3">📄 الصق نص قائمة الامتحان هنا (بصيغة Markdown)</label>
                    <textarea id="mdContent" name="content" required 
                        class="w-full min-h-80 p-4 border-2 border-gray-300 rounded-lg text-base focus:outline-none focus:border-green-700 transition resize-vertical"
                        placeholder="الصق النص هنا...

مثال:
## غرداية الفـــــرع:
## المنيعة الولايــــة:
## 0000/ثانوية ........ - المنيعة مركز الإجراء:
# 94748 لجنة التلميذ
## 1666 .... 15-05-2010 أحمد محمد بن علي 226 1"></textarea>
                </div>
                
                <!-- Options -->
                <div class="bg-gray-100 p-5 rounded-lg">
                    <h3 class="font-semibold text-gray-800 mb-4">⚙️ خيارات التصدير</h3>
                    <div class="flex flex-wrap gap-8">
                        <label class="flex items-center cursor-pointer font-normal">
                            <input type="checkbox" name="single_sheet" value="1" checked class="w-5 h-5 ml-2 cursor-pointer">
                            ملف واحد (جميع التلاميذ في ورقة واحدة)
                        </label>
                        <label class="flex items-center cursor-pointer font-normal">
                            <input type="checkbox" name="multi_sheet" value="1" checked class="w-5 h-5 ml-2 cursor-pointer">
                            ملف متعدد (ورقة لكل لجنة)
                        </label>
                    </div>
                </div>
                
                <!-- Buttons -->
                <div class="flex gap-3">
                    <button type="submit" class="flex-1 bg-gradient-to-br from-green-700 to-green-900 text-white font-bold py-4 px-6 rounded-lg hover:shadow-lg hover:-translate-y-0.5 transition">
                        📥 تحويل وتنزيل
                    </button>
                    <button type="button" onclick="clearForm()" class="flex-1 bg-gray-500 text-white font-bold py-4 px-6 rounded-lg hover:bg-gray-600 hover:shadow-lg hover:-translate-y-0.5 transition">
                        🗑️ مسح
                    </button>
                </div>
            </form>
            
            <!-- Loading Spinner -->
            <div id="loading" class="hidden text-center py-5">
                <div class="spinner border-4 border-gray-300 border-t-green-700 rounded-full w-12 h-12 mx-auto mb-4"></div>
                <p>جاري المعالجة...</p>
            </div>
            
            <!-- Messages -->
            <div id="message" class="hidden p-4 rounded-lg mt-5"></div>
            
            <!-- Example -->
            <div class="bg-gray-100 border-r-4 border-red-700 p-4 rounded my-5 font-mono text-sm">
                <div class="font-semibold mb-2 text-gray-800 font-sans">💡 مثال على التنسيق المطلوب:</div>
                <div class="text-gray-700">## غرداية الفـــــرع:</div>
                <div class="text-gray-700">## .... الولايــــة:</div>
                <div class="text-gray-700">## 58003/ثانوية ..... - المنيعة مركز الإجراء:</div>
                <div class="text-gray-700"># 94748 لجنة التلميذ</div>
                <div class="text-gray-700">## 1666 .... 15-05-2010 أحمد محمد بن علي 222 1</div>
            </div>
        </div>
    </div>
    
    <script>
        document.getElementById('converterForm').addEventListener('submit', async function(e) {
            e.preventDefault();
            
            const content = document.getElementById('mdContent').value;
            const singleSheet = document.querySelector('input[name="single_sheet"]').checked;
            const multiSheet = document.querySelector('input[name="multi_sheet"]').checked;
            const loading = document.getElementById('loading');
            const message = document.getElementById('message');
            
            message.classList.add('hidden');
            
            if (!singleSheet && !multiSheet) {
                showMessage('الرجاء اختيار نوع واحد على الأقل من أنواع التصدير', 'error');
                return;
            }
            
            loading.classList.remove('hidden');
            
            try {
                const formData = new FormData();
                formData.append('content', content);
                formData.append('single_sheet', singleSheet ? '1' : '0');
                formData.append('multi_sheet', multiSheet ? '1' : '0');
                
                const response = await fetch('/convert', {
                    method: 'POST',
                    body: formData
                });
                
                if (!response.ok) {
                    const error = await response.json();
                    throw new Error(error.error || 'خطأ في المعالجة');
                }
                
                const blob = await response.blob();
                const filename = response.headers.get('Content-Disposition')
                    .split('filename=')[1]
                    .replace(/"/g, '');
                
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = filename;
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                document.body.removeChild(a);
                
                showMessage('✅ تم التحويل والتنزيل بنجاح!', 'success');
            } catch (error) {
                showMessage('❌ ' + error.message, 'error');
            } finally {
                loading.classList.add('hidden');
            }
        });
        
        function showMessage(text, type) {
            const message = document.getElementById('message');
            message.textContent = text;
            message.classList.remove('hidden');
            if (type === 'success') {
                message.className = 'hidden p-4 rounded-lg mt-5 bg-green-100 border border-green-400 text-green-800';
            } else {
                message.className = 'hidden p-4 rounded-lg mt-5 bg-red-100 border border-red-400 text-red-900';
            }
            message.classList.remove('hidden');
        }
        
        function clearForm() {
            document.getElementById('mdContent').value = '';
            document.getElementById('message').classList.add('hidden');
        }
    </script>
</body>
</html>
"""


def create_web_app():
    """Create and configure Flask web application."""
    app = Flask(__name__)
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

    @app.route('/')
    def index():
        return render_template_string(HTML_TEMPLATE)

    @app.route('/convert', methods=['POST'])
    def convert():
        try:
            content = request.form.get('content', '').strip()
            single_sheet = request.form.get('single_sheet') == '1'
            multi_sheet = request.form.get('multi_sheet') == '1'

            if not content:
                return jsonify({'error': 'لم يتم إدخال أي محتوى'}), 400

            # Parse content
            records = parse_md_content(content)

            if not records:
                return jsonify({'error': 'لم يتم العثور على سجلات صالحة. تأكد من صحة التنسيق.'}), 400

            # Get center info for filename
            center_info = records[0].get('center_info', 'exam_list')
            # Clean filename - remove special characters
            clean_name = re.sub(r'[^\u0600-\u06FF\w\s]', '_', center_info)
            clean_name = re.sub(r'\s+', '_', clean_name)

            # Create Excel files based on options
            if not single_sheet and not multi_sheet:
                return jsonify({'error': 'يجب اختيار نوع واحد على الأقل'}), 400

            # If only one type selected, return that single file
            if single_sheet and not multi_sheet:
                output = create_single_sheet_excel(records)
                filename = f"{clean_name}_single.xlsx"
                output.seek(0)
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
            
            elif multi_sheet and not single_sheet:
                output = create_multi_sheet_excel(records)
                filename = f"{clean_name}_committees.xlsx"
                output.seek(0)
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
            
            else:
                # Both selected - create both files and return as ZIP
                import zipfile
            
            single_output = create_single_sheet_excel(records)
            multi_output = create_multi_sheet_excel(records)
            
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                single_output.seek(0)
                multi_output.seek(0)
                zip_file.writestr(f"{clean_name}_single.xlsx", single_output.read())
                zip_file.writestr(f"{clean_name}_committees.xlsx", multi_output.read())
            
            zip_buffer.seek(0)
            return send_file(
                zip_buffer,
                mimetype='application/zip',
                as_attachment=True,
                download_name=f"{clean_name}_excel_files.zip"
            )

        except Exception as e:
            return jsonify({'error': f'خطأ في المعالجة: {str(e)}'}), 500

    return app


def cli_mode():
    """Command-line interface mode."""
    parser = argparse.ArgumentParser(
        description='Convert Algerian exam lists from Markdown to Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python md_to_excel_web.py --cli input.md                    # Creates both files in current directory
  python md_to_excel_web.py --cli input.md -o results/output  # Custom output path
  python md_to_excel_web.py --cli input.md --single-only      # Only single-sheet
  python md_to_excel_web.py --cli input.md --multi-only       # Only multi-sheet
        """
    )

    parser.add_argument('--cli', dest='input_file', help='Input Markdown file path')
    parser.add_argument('-o', '--output', help='Output path/name (without extension)')
    parser.add_argument('--single-only', action='store_true', help='Only create single-sheet file')
    parser.add_argument('--multi-only', action='store_true', help='Only create multi-sheet file')
    parser.add_argument('--port', type=int, default=5000, help='Web server port (default: 5000)')

    args = parser.parse_args()

    if args.input_file:
        # CLI mode
        input_path = Path(args.input_file)
        if not input_path.exists():
            print(f"[ERROR] File not found: {input_path}")
            sys.exit(1)

        # Parse the file
        print(f"[*] Parsing {input_path}...")
        records = parse_md_file(str(input_path))

        if not records:
            print("[ERROR] No records found in the file.")
            print("        Make sure the file follows the expected format:")
            print("        ## 47112001 [birthplace] DD-MM-YYYY [name] [surname] 2947XXXX [order]")
            sys.exit(1)

        print(f"[*] Found {len(records)} student records")

        # Get center info for filename
        center_info = records[0].get('center_info', 'exam_list')
        clean_name = re.sub(r'[^\u0600-\u06FF\w\s]', '_', center_info)
        clean_name = re.sub(r'\s+', '_', clean_name)

        # Determine output path
        if args.output:
            output_path = Path(args.output)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            base_name = str(output_path)
        else:
            base_name = str(Path.cwd() / clean_name)

        # Generate files based on options
        if not args.multi_only:
            single_output = f"{base_name}_single.xlsx"
            create_single_sheet_excel(records, single_output)

        if not args.single_only:
            multi_output = f"{base_name}_committees.xlsx"
            create_multi_sheet_excel(records, multi_output)

        print("[*] Done!")
    else:
        # Web mode
        if not FLASK_AVAILABLE:
            print("[ERROR] Flask is required for web mode. Install it with: pip install flask")
            sys.exit(1)

        print("=" * 60)
        print("محول قوائم الامتحانات - وزارة التربية الوطنية")
        print("Algerian Exam List Converter - Web Interface")
        print("=" * 60)
        print(f"\n🌐 Starting web server on http://localhost:{args.port}")
        print(f"📝 Open your browser and navigate to: http://localhost:{args.port}")
        print("\nPress Ctrl+C to stop the server\n")

        app = create_web_app()
        app.run(host='0.0.0.0', port=args.port, debug=False)


if __name__ == "__main__":
    cli_mode()